"""
Finance export routes for financial data export functionality
Handles CSV and Excel exports for proposal summaries and client reports
"""
from flask import Blueprint, request, jsonify, send_file
from typing import Optional, Dict, Any, List
from datetime import datetime, timedelta
import csv
import io
import os
import tempfile
from api.utils.decorators import token_required, finance_required
from api.utils.database import get_db_connection
import psycopg2.extras
import json

bp = Blueprint('finance_export', __name__)


def _ensure_str(v):
    if v is None:
        return ''
    return str(v)


def _rows_to_pdf_bytes(title: str, headers: List[str], rows: List[List[Any]]):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except Exception:
        raise RuntimeError('PDF export not available (reportlab missing)')

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    y = height - 40
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, title[:120])
    y -= 18

    c.setFont('Helvetica', 8)
    c.drawString(40, y, 'Generated: ' + datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'))
    y -= 18

    c.setFont('Helvetica-Bold', 8)
    c.drawString(40, y, ' | '.join([h[:28] for h in headers])[:140])
    y -= 14
    c.setFont('Helvetica', 8)

    for r in rows:
        line = ' | '.join([_ensure_str(v) for v in r])
        c.drawString(40, y, line[:140])
        y -= 12
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont('Helvetica', 8)

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.getvalue()

def _extract_amount_from_content(content_data):
    """Extract financial amount from proposal content using same logic as frontend"""
    if not content_data:
        return 0.0
    
    def _parse_num(v):
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return float(v)
        cleaned = str(v).replace(',', '').replace('R', '').replace('$', '').strip()
        try:
            return float(cleaned)
        except:
            return 0.0
    
    # Try direct amount fields first
    if isinstance(content_data, dict):
        for key in ['budget', 'amount', 'total', 'value', 'price']:
            if key in content_data:
                amount = _parse_num(content_data[key])
                if amount > 0:
                    return amount
    
    # Try to extract from pricing tables in sections
    if isinstance(content_data, dict) and 'sections' in content_data:
        total = 0.0
        for section in content_data.get('sections', []):
            if not isinstance(section, dict):
                continue
                
            # Check for tables in section
            for table in section.get('tables', []):
                if isinstance(table, dict) and table.get('type') == 'price':
                    cells = table.get('cells', [])
                    if isinstance(cells, list) and len(cells) > 1:
                        # Look for total column
                        header_row = cells[0] if isinstance(cells[0], list) else []
                        total_col_idx = None
                        
                        for i, header in enumerate(header_row):
                            if isinstance(header, str) and 'total' in header.lower():
                                total_col_idx = i
                                break
                        
                        if total_col_idx is not None:
                            for row in cells[1:]:
                                if isinstance(row, list) and len(row) > total_col_idx:
                                    total += _parse_num(row[total_col_idx])
        
        return total
    
    return 0.0

def _format_currency(amount):
    """Format amount as South African Rand"""
    if amount <= 0:
        return "R 0"
    return f"R {amount:,.2f}"


def _get_proposals_with_financials(user_id=None, status_filter=None, date_from=None, date_to=None):
    """Get proposals with financial calculations. Uses dynamic column names to support both owner_id/user_id and client/client_name."""
    proposals = []

    params = []

    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            cursor.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'proposals'
                """
            )
            existing_columns = [row['column_name'] for row in cursor.fetchall()]

            cursor.execute(
                """
                SELECT EXISTS(
                    SELECT 1
                    FROM information_schema.tables
                    WHERE table_name = 'users'
                ) AS users_table_exists
                """
            )
            users_table_exists = bool((cursor.fetchone() or {}).get('users_table_exists'))

            users_full_name_exists = False
            if users_table_exists:
                cursor.execute(
                    """
                    SELECT EXISTS(
                        SELECT 1
                        FROM information_schema.columns
                        WHERE table_name = 'users' AND column_name = 'full_name'
                    ) AS full_name_exists
                    """
                )
                users_full_name_exists = bool((cursor.fetchone() or {}).get('full_name_exists'))

            owner_col = None
            if 'created_by' in existing_columns:
                owner_col = 'created_by'
            elif 'owner_id' in existing_columns:
                owner_col = 'owner_id'
            elif 'user_id' in existing_columns:
                owner_col = 'user_id'

            client_col = None
            if 'client_name' in existing_columns:
                client_col = 'client_name'
            elif 'client' in existing_columns:
                client_col = 'client'

            content_col = None
            if 'content' in existing_columns:
                content_col = 'content'
            elif 'sections' in existing_columns:
                content_col = 'sections'

            select_cols = [
                'p.id',
                'p.title',
                'p.status',
            ]
            if client_col:
                select_cols.append(f"p.{client_col} AS client_name")
            else:
                select_cols.append("NULL AS client_name")

            if 'created_at' in existing_columns:
                select_cols.append('p.created_at')
            else:
                select_cols.append('NULL AS created_at')

            if 'updated_at' in existing_columns:
                select_cols.append('p.updated_at')
            else:
                select_cols.append('NULL AS updated_at')

            if content_col:
                select_cols.append(f"p.{content_col} AS content")
            else:
                select_cols.append("NULL AS content")

            join_sql = ""
            if owner_col and users_table_exists and users_full_name_exists:
                join_sql = f"LEFT JOIN users u ON u.id::text = p.{owner_col}::text"
                select_cols.append("u.full_name AS created_by")
            elif owner_col:
                select_cols.append(f"p.{owner_col}::text AS created_by")
            else:
                select_cols.append("NULL AS created_by")

            query = f"""
                SELECT {', '.join(select_cols)}
                FROM proposals p
                {join_sql}
                WHERE 1=1
            """

            if status_filter:
                query += " AND LOWER(p.status) LIKE %s"
                params.append(f"%{status_filter.lower()}%")

            if date_from and 'created_at' in existing_columns:
                query += " AND p.created_at >= %s"
                params.append(date_from)

            if date_to and 'created_at' in existing_columns:
                query += " AND p.created_at <= %s"
                params.append(date_to)

            if 'created_at' in existing_columns:
                query += " ORDER BY p.created_at DESC"
            else:
                query += " ORDER BY p.id DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall() or []
            
            for row in rows:
                amount = 0.0
                if row.get('content'):
                    try:
                        content_data = json.loads(row['content']) if isinstance(row['content'], str) else row['content']
                        amount = _extract_amount_from_content(content_data)
                    except Exception:
                        pass
                
                days_in_status = 0
                if row.get('updated_at'):
                    try:
                        days_in_status = (datetime.now() - row['updated_at']).days
                    except Exception:
                        pass
                
                proposal = {
                    'id': row['id'],
                    'title': row.get('title') or '',
                    'client_name': row.get('client_name') or '',
                    'status': row.get('status') or '',
                    'created_at': row.get('created_at'),
                    'updated_at': row.get('updated_at'),
                    'created_by': row.get('created_by') or '',
                    'amount': amount,
                    'formatted_amount': _format_currency(amount),
                    'days_in_status': days_in_status
                }
                proposals.append(proposal)
                
    except Exception as e:
        print(f"Error fetching proposals: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return proposals

def _get_client_portfolio_data(proposals):
    """Aggregate proposal data by client"""
    client_data = {}
    
    for proposal in proposals:
        client = proposal['client_name'] or 'Unknown Client'
        
        if client not in client_data:
            client_data[client] = {
                'client_name': client,
                'total_proposals': 0,
                'total_amount': 0.0,
                'approved_proposals': 0,
                'approved_amount': 0.0,
                'pending_proposals': 0,
                'pending_amount': 0.0,
                'proposals': []
            }
        
        client_data[client]['total_proposals'] += 1
        client_data[client]['total_amount'] += proposal['amount']
        client_data[client]['proposals'].append(proposal)
        
        status = proposal['status'].lower()
        if 'approved' in status or 'signed' in status:
            client_data[client]['approved_proposals'] += 1
            client_data[client]['approved_amount'] += proposal['amount']
        elif 'pending' in status or 'review' in status:
            client_data[client]['pending_proposals'] += 1
            client_data[client]['pending_amount'] += proposal['amount']
    
    # Calculate additional metrics
    for client in client_data.values():
        client['success_rate'] = (client['approved_proposals'] / client['total_proposals'] * 100) if client['total_proposals'] > 0 else 0
        client['average_deal_size'] = client['total_amount'] / client['total_proposals'] if client['total_proposals'] > 0 else 0
        client['formatted_total'] = _format_currency(client['total_amount'])
        client['formatted_average'] = _format_currency(client['average_deal_size'])
    
    return list(client_data.values())

def _generate_csv_summary(proposals):
    """Generate CSV for proposal financial summary (Excel-friendly: UTF-8 BOM, CRLF)."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator='\r\n')
    
    # Header
    writer.writerow([
        'Proposal ID', 'Title', 'Client', 'Status', 
        'Created Date', 'Updated Date', 'Amount (ZAR)', 
        'Days in Status', 'Created By'
    ])
    
    # Data rows
    for proposal in proposals:
        writer.writerow([
            proposal['id'],
            proposal['title'],
            proposal['client_name'],
            proposal['status'],
            proposal['created_at'].strftime('%Y-%m-%d') if proposal['created_at'] else '',
            proposal['updated_at'].strftime('%Y-%m-%d') if proposal['updated_at'] else '',
            f"{proposal['amount']:.2f}",
            proposal['days_in_status'],
            proposal['created_by']
        ])
    
    return output.getvalue()

def _generate_csv_client_report(client_data):
    """Generate CSV for client financial report (Excel-friendly: UTF-8 BOM, CRLF)."""
    output = io.StringIO()
    writer = csv.writer(output, lineterminator='\r\n')
    
    # Header
    writer.writerow([
        'Client Name', 'Total Proposals', 'Total Amount (ZAR)', 
        'Approved Proposals', 'Approved Amount (ZAR)', 
        'Pending Proposals', 'Pending Amount (ZAR)',
        'Success Rate (%)', 'Average Deal Size (ZAR)'
    ])
    
    # Data rows
    for client in client_data:
        writer.writerow([
            client['client_name'],
            client['total_proposals'],
            f"{client['total_amount']:.2f}",
            client['approved_proposals'],
            f"{client['approved_amount']:.2f}",
            client['pending_proposals'],
            f"{client['pending_amount']:.2f}",
            f"{client['success_rate']:.1f}",
            f"{client['average_deal_size']:.2f}"
        ])
    
    return output.getvalue()

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _generate_xlsx_summary(proposals):
    """Generate Excel (.xlsx) for proposal financial summary."""
    if not OPENPYXL_AVAILABLE:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal Summary"
    headers = [
        'Proposal ID', 'Title', 'Client', 'Status',
        'Created Date', 'Updated Date', 'Amount (ZAR)',
        'Days in Status', 'Created By'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, proposal in enumerate(proposals, 2):
        ws.cell(row=row_idx, column=1, value=proposal['id'])
        ws.cell(row=row_idx, column=2, value=proposal['title'])
        ws.cell(row=row_idx, column=3, value=proposal['client_name'])
        ws.cell(row=row_idx, column=4, value=proposal['status'])
        ws.cell(row=row_idx, column=5,
               value=proposal['created_at'].strftime('%Y-%m-%d') if proposal['created_at'] else '')
        ws.cell(row=row_idx, column=6,
               value=proposal['updated_at'].strftime('%Y-%m-%d') if proposal['updated_at'] else '')
        ws.cell(row=row_idx, column=7, value=proposal['amount'])
        ws.cell(row=row_idx, column=8, value=proposal['days_in_status'])
        ws.cell(row=row_idx, column=9, value=proposal['created_by'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_xlsx_client_report(client_data):
    """Generate Excel (.xlsx) for client financial report."""
    if not OPENPYXL_AVAILABLE:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Report"
    headers = [
        'Client Name', 'Total Proposals', 'Total Amount (ZAR)',
        'Approved Proposals', 'Approved Amount (ZAR)',
        'Pending Proposals', 'Pending Amount (ZAR)',
        'Success Rate (%)', 'Average Deal Size (ZAR)'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, client in enumerate(client_data, 2):
        ws.cell(row=row_idx, column=1, value=client['client_name'])
        ws.cell(row=row_idx, column=2, value=client['total_proposals'])
        ws.cell(row=row_idx, column=3, value=client['total_amount'])
        ws.cell(row=row_idx, column=4, value=client['approved_proposals'])
        ws.cell(row=row_idx, column=5, value=client['approved_amount'])
        ws.cell(row=row_idx, column=6, value=client['pending_proposals'])
        ws.cell(row=row_idx, column=7, value=client['pending_amount'])
        ws.cell(row=row_idx, column=8, value=client['success_rate'])
        ws.cell(row=row_idx, column=9, value=client['average_deal_size'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route("/finance/export/proposal-summary", methods=["GET", "OPTIONS"])
@token_required
def export_proposal_summary(username=None, user_id=None, email=None):
    """Export proposal financial summary"""
    try:
        # Get query parameters
        format_type = request.args.get('format', 'csv').lower()
        status_filter = request.args.get('status')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Get proposals data
        proposals = _get_proposals_with_financials(
            user_id=user_id,
            status_filter=status_filter,
            date_from=date_from,
            date_to=date_to
        )

        if not proposals:
            return jsonify({
                'error': 'No proposals found to export',
                'report': 'proposal_summary',
                'format': format_type,
            }), 404
        
        ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')

        if format_type == 'csv':
            csv_data = _generate_csv_summary(proposals)
            
            # Excel-friendly: UTF-8 BOM so Excel parses columns correctly
            output = io.BytesIO()
            output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
            output.write(csv_data.encode('utf-8'))
            output.seek(0)
            
            filename = f"proposal_summary_{ts}.csv"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='text/csv'
            )
        
        elif format_type == 'pdf':
            headers = [
                'Proposal ID', 'Title', 'Client', 'Status',
                'Created Date', 'Updated Date', 'Amount (ZAR)',
                'Days in Status', 'Created By'
            ]

            rows = []
            for proposal in proposals:
                rows.append([
                    proposal.get('id'),
                    (proposal.get('title') or '')[:60],
                    (proposal.get('client_name') or '')[:40],
                    (proposal.get('status') or '')[:30],
                    proposal.get('created_at').strftime('%Y-%m-%d') if proposal.get('created_at') else '',
                    proposal.get('updated_at').strftime('%Y-%m-%d') if proposal.get('updated_at') else '',
                    f"{float(proposal.get('amount') or 0.0):.2f}",
                    proposal.get('days_in_status'),
                    (proposal.get('created_by') or '')[:40],
                ])

            pdf_bytes = _rows_to_pdf_bytes('Proposal Financial Summary', headers, rows)
            output = io.BytesIO(pdf_bytes)
            output.seek(0)
            filename = f"proposal_summary_{ts}.pdf"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )

        else:
            return jsonify({'error': 'Only CSV and PDF formats are currently supported'}), 400
            
    except Exception as e:
        print(f"Error exporting proposal summary: {e}")
        return jsonify({'error': 'Failed to generate export'}), 500

@bp.route("/finance/export/client-report", methods=["GET", "OPTIONS"])
@token_required
def export_client_report(username=None, user_id=None, email=None):
    """Export client financial report"""
    try:
        # Get query parameters
        format_type = request.args.get('format', 'csv').lower()
        status_filter = request.args.get('status')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Get proposals data
        proposals = _get_proposals_with_financials(
            user_id=user_id,
            status_filter=status_filter,
            date_from=date_from,
            date_to=date_to
        )

        if not proposals:
            return jsonify({
                'error': 'No proposals found to export',
                'report': 'client_report',
                'format': format_type,
            }), 404
        
        # Aggregate by client
        client_data = _get_client_portfolio_data(proposals)
        
        ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')

        if format_type == 'csv':
            csv_data = _generate_csv_client_report(client_data)
            
            # Excel-friendly: UTF-8 BOM so Excel parses columns correctly
            output = io.BytesIO()
            output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
            output.write(csv_data.encode('utf-8'))
            output.seek(0)
            
            filename = f"client_report_{ts}.csv"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='text/csv'
            )

        elif format_type == 'pdf':
            headers = [
                'Client Name', 'Total Proposals', 'Total Amount (ZAR)',
                'Approved Proposals', 'Approved Amount (ZAR)',
                'Pending Proposals', 'Pending Amount (ZAR)',
                'Success Rate (%)', 'Average Deal Size (ZAR)'
            ]

            rows = []
            for client in client_data:
                rows.append([
                    (client.get('client_name') or '')[:50],
                    client.get('total_proposals'),
                    f"{float(client.get('total_amount') or 0.0):.2f}",
                    client.get('approved_proposals'),
                    f"{float(client.get('approved_amount') or 0.0):.2f}",
                    client.get('pending_proposals'),
                    f"{float(client.get('pending_amount') or 0.0):.2f}",
                    f"{float(client.get('success_rate') or 0.0):.1f}",
                    f"{float(client.get('average_deal_size') or 0.0):.2f}",
                ])

            pdf_bytes = _rows_to_pdf_bytes('Client Financial Report', headers, rows)
            output = io.BytesIO(pdf_bytes)
            output.seek(0)
            filename = f"client_report_{ts}.pdf"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )

        else:
            return jsonify({'error': 'Only CSV and PDF formats are currently supported'}), 400
            
    except Exception as e:
        print(f"Error exporting client report: {e}")
        return jsonify({'error': 'Failed to generate export'}), 500

@bp.route("/finance/export/summary-stats", methods=["GET", "OPTIONS"])
@token_required
def get_export_summary_stats(username=None, user_id=None, email=None):
    """Get summary statistics for export dialog"""
    try:
        # Get basic counts
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            
            # Total proposals
            cursor.execute("SELECT COUNT(*) as total FROM proposals")
            total_proposals = cursor.fetchone()['total']
            
            # Proposals by status
            cursor.execute("""
                SELECT status, COUNT(*) as count 
                FROM proposals 
                GROUP BY status 
                ORDER BY count DESC
            """)
            status_breakdown = cursor.fetchall()
            
            # Date range
            cursor.execute("""
                SELECT MIN(created_at) as earliest, MAX(created_at) as latest 
                FROM proposals
            """)
            date_range = cursor.fetchone()
        
        return jsonify({
            'total_proposals': total_proposals,
            'status_breakdown': status_breakdown,
            'date_range': date_range,
            'supported_formats': ['csv', 'pdf'],
            'report_types': ['proposal_summary', 'client_report']
        })
        
    except Exception as e:
        print(f"Error getting export stats: {e}")
        return jsonify({'error': 'Failed to get export statistics'}), 500
